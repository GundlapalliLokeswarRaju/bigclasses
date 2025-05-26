# enrollments/views.py
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
from django.conf import settings
import logging
logger = logging.getLogger(__name__)
class EnrollmentView(APIView):
    EXCEL_HEADERS = ['Timestamp', 'Student Name', 'Email', 'Course', 'Phone']
    def post(self, request):
        try:
            required_fields = ['student_name', 'email', 'course_title', 'phone']
            if not all(field in request.data for field in required_fields):
                return Response({'error': 'Missing required fields'}, status=status.HTTP_400_BAD_REQUEST)

            # Always define these variables
            media_root = os.path.join(settings.BASE_DIR, 'media')
            enrollments_dir = os.path.join(media_root, 'enrollments')
            file_path = os.path.join(enrollments_dir, 'enrollments.xlsx')
            os.makedirs(enrollments_dir, exist_ok=True)

            try:
                wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
                ws = wb.active

                if ws['A1'].value != self.EXCEL_HEADERS[0]:
                    ws.delete_rows(1, ws.max_row)                    
                    ws.append(self.EXCEL_HEADERS)

                enrollment_data = [
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    request.data.get('student_name', '').strip(),
                    request.data.get('email', '').strip().lower(),
                    request.data.get('course_title', '').strip(),
                    request.data.get('phone', '').strip()
                ]

                ws.append(enrollment_data)

                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(file_path)
                wb.close()
                    'data': dict(zip(self.EXCEL_HEADERS, enrollment_data))
                }, status=status.HTTP_201_CREATED)
                    'message': 'Enrollment successful',
            except Exception as e:ip(self.EXCEL_HEADERS, enrollment_data))
                logger.error(f"Excel handling error: {str(e)}")
                raise
            except Exception as e:
        except Exception as e:"Excel handling error: {str(e)}")
            logger.error(f"Enrollment error: {str(e)}")
            return Response({
                'error': f'Enrollment failed: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            return Response({
                'error': f'Enrollment failed: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
